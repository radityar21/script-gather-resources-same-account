##### Need to set environment variables with below

# FORMAT --> Key = Value
# LOOKBACK_DAYS =30
# REPORT_S3_BUCKET =cur-bucket-ap-southeast-3
# REPORT_S3_PREFIX = lambda-reports

import boto3
import os
import json
import datetime
import csv
import logging
import io
from botocore.config import Config
from botocore.exceptions import ClientError
import openpyxl
from openpyxl.styles import Font
from openpyxl import Workbook


from io import BytesIO

# Logging
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# Env vars
LOOKBACK_DAYS = int(os.environ.get('LOOKBACK_DAYS', '30'))
REGION = os.environ.get('AWS_REGION', 'ap-southeast-3')
# REPORT_S3_BUCKET = os.environ.get('REPORT_S3_BUCKET')
OUTPUT_BUCKET = os.environ.get('REPORT_S3_BUCKET')
REPORT_S3_PREFIX = os.environ.get('REPORT_S3_PREFIX', 'lambda-reports')


# Initialize boto3 clients
ec2 = boto3.client("ec2")
s3 = boto3.client("s3")
efs = boto3.client("efs")
elb = boto3.client("elbv2")
elc = boto3.client("elasticache")
ce = boto3.client("ce", region_name="us-east-1")  # CE must be in us-east-1
sts = boto3.client("sts")
boto_config = Config(retries={'max_attempts': 5})

# Global prefix for this run
RUN_PREFIX = datetime.datetime.utcnow().strftime("%Y%m%d-%H%M%S")


def human_readable_size(size_in_bytes):
    if size_in_bytes is None or size_in_bytes == 0:
        return "-"
    units = ["Bytes", "KB", "MB", "GB", "TB"]
    size = float(size_in_bytes)
    for unit in units:
        if size < 1024.0:
            return f"{size:.2f} {unit}"
        size /= 1024.0
    return f"{size:.2f} PB"

def collect_ec2():
    sheet_name = "EC2 Instances"
    headers = ["#", "Name", "Instance ID", "Instance state", "Instance type",
               "Elastic IP", "Launch time", "vCPUs", "Memory (GiB)", "Disk GiB",
               "Average CPU%", "Average Memory%"]

    instances = []
    reservations = ec2.describe_instances()["Reservations"]
    for idx, res in enumerate(reservations, 1):
        for inst in res["Instances"]:
            name = next((tag["Value"] for tag in inst.get("Tags", []) if tag["Key"] == "Name"), "-")
            elastic_ip = "-"
            try:
                addresses = ec2.describe_addresses(Filters=[{"Name": "instance-id", "Values": [inst["InstanceId"]]}])
                if addresses.get("Addresses"):
                    elastic_ip = addresses["Addresses"][0].get("PublicIp", "-")
            except ClientError:
                pass

            instances.append([
                idx,
                name,
                inst["InstanceId"],
                inst["State"]["Name"],
                inst["InstanceType"],
                elastic_ip,
                str(inst["LaunchTime"]),
                "-", "-", "-", "-", "-"
            ])
    return sheet_name, headers, instances

def collect_s3():
    sheet_name = "Simple Storage Service"
    headers = ["#", "Name", "Region", "Creation Date", "Size"]

    response = s3.list_buckets()
    results = []
    for idx, bucket in enumerate(response["Buckets"], 1):
        try:
            region = s3.get_bucket_location(Bucket=bucket["Name"])["LocationConstraint"]
            if not region:
                region = "us-east-1"
        except Exception:
            region = "-"

        # Calculate total size
        total_size = 0
        try:
            s3_res = boto3.resource("s3")
            bucket_obj = s3_res.Bucket(bucket["Name"])
            for obj in bucket_obj.objects.all():
                total_size += obj.size
            size_str = human_readable_size(total_size)
        except Exception:
            size_str = "-"

        results.append([
            idx,
            bucket["Name"],
            region,
            str(bucket["CreationDate"]),
            size_str
        ])
    return sheet_name, headers, results

def collect_elc():
    sheet_name = "Elasticache List"
    headers = ["#", "Name", "Node types", "Types"]

    results = []
    clusters = elc.describe_cache_clusters(ShowCacheNodeInfo=True)["CacheClusters"]
    for idx, cl in enumerate(clusters, 1):
        results.append([idx, cl["CacheClusterId"], cl["CacheNodeType"], cl["Engine"]])
        for i, node in enumerate(cl["CacheNodes"], 1):
            results.append([f"{idx}.{i}", f"{cl['CacheClusterId']}-{node['CacheNodeId']}", "", f"Node {node['CacheNodeStatus']}"])
    return sheet_name, headers, results

def collect_elb():
    sheet_name = "Elastic Load Balancer"
    headers = ["#", "Name", "State", "Type", "Scheme", "IP address type", "VPC ID", "Security groups", "Date created", "DNS name"]

    results = []
    lbs = elb.describe_load_balancers()["LoadBalancers"]
    for idx, lb in enumerate(lbs, 1):
        results.append([
            idx,
            lb["LoadBalancerName"],
            lb["State"]["Code"],
            lb["Type"],
            lb["Scheme"],
            lb.get("IpAddressType", "-"),
            lb["VpcId"],
            ", ".join(lb.get("SecurityGroups", [])),
            str(lb["CreatedTime"]),
            lb["DNSName"]
        ])
    return sheet_name, headers, results

def collect_efs():
    sheet_name = "Elastic File System"
    headers = ["#", "Name", "File system ID", "Encrypted", "Total size", "Size in EFS Standard",
               "Size in EFS IA", "Size in Archive", "File system state", "Creation time"]

    results = []
    filesystems = efs.describe_file_systems()["FileSystems"]
    for idx, fs in enumerate(filesystems, 1):
        results.append([
            idx,
            fs.get("Name", "-"),
            fs["FileSystemId"],
            "Encrypted" if fs["Encrypted"] else "Unencrypted",
            human_readable_size(fs.get("SizeInBytes", {}).get("Value")),
            "-", "-", "-",
            fs["LifeCycleState"],
            str(fs["CreationTime"])
        ])
    return sheet_name, headers, results

def collect_vpc():
    sheet_name = "Virtual Private Cloud"
    headers = ["#", "Services", "Qty"]

    results = [
        [1, "AWS Site-to-site VPN", "1"],
        [2, "Idle public IPv4 address", "2"],
        [3, "In-use public IPv4 address", "4"],
    ]
    return sheet_name, headers, results


def lambda_handler(event, context):
    account_id = sts.get_caller_identity()["Account"]
    timestamp = datetime.datetime.utcnow().strftime("%Y-%m-%d_%H-%M-%S")
    prefix = f"reports/{account_id}/{timestamp}/"

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Account-{account_id}"[:31]  # One sheet per account

    collectors = [
        collect_ec2,
        collect_s3,
        collect_elc,
        collect_elb,
        collect_efs,
        collect_vpc,
    ]

    row_cursor = 1
    for collector in collectors:
        sheet_name, headers, rows = collector()

        # Log count to CloudWatch
        logger.info(f"{sheet_name} gathered: {len(rows)}")

        # Section Title
        ws.cell(row=row_cursor, column=1, value=sheet_name).font = Font(bold=True, size=12)
        row_cursor += 1

        # Headers
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_cursor, column=col_idx, value=header).font = Font(bold=True)
        row_cursor += 1

        # Rows
        for row in rows:
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_cursor, column=col_idx, value=val)
            row_cursor += 1

        # Summary Count inside Excel
        ws.cell(row=row_cursor, column=1, value=f"Total {sheet_name}: {len(rows)}").font = Font(bold=True, italic=True)
        row_cursor += 2  # Add spacing before next section

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    key = f"{prefix}aws_inventory_report.xlsx"
    s3.put_object(Bucket=OUTPUT_BUCKET, Key=key, Body=output.getvalue())

    logger.info(f"Final report saved: s3://{OUTPUT_BUCKET}/{key}")

    return {
        "status": "success",
        "s3_key": key,
        "bucket": OUTPUT_BUCKET,
        "account": account_id
    }
